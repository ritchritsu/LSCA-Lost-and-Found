import psutil  # Add to requirements.txt
import time
from datetime import datetime
import threading
import csv
from pathlib import Path
import os
import statistics
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class SystemMonitor:
    def __init__(self, interval=1):
        """Initialize system monitor"""
        try:
            self.interval = interval
            self.monitoring = False
            self.metrics = []
            self.start_time = datetime.now()
            
            # Initialize disk I/O tracking with absolute values
            initial_io = psutil.disk_io_counters()
            self.initial_read_bytes = initial_io.read_bytes
            self.initial_write_bytes = initial_io.write_bytes
            self.last_io = initial_io
            self.total_read_bytes = initial_io.read_bytes  # Changed: Start with current value
            self.total_write_bytes = initial_io.write_bytes  # Changed: Start with current value
            
            # Initialize CPU tracking with better defaults
            self.cpu_values = []
            self.peak_metrics = {
                'cpu': 0,
                'cpu_low': float('inf'),  # Changed: Start with infinity for proper min tracking
                'memory': 0,
                'disk_read': 0,
                'disk_write': 0
            }
        except Exception as e:
            print(f"Error initializing monitor: {e}")
            self.monitoring = False
    
    def collect_metrics(self):
        """Collect system metrics"""
        try:
            # Get CPU metrics and round to 2 decimal places
            cpu_percent = psutil.cpu_percent(interval=1.0)
            cpu_percent = round(cpu_percent, 2)
            
            # Update CPU tracking
            if cpu_percent > 0:
                self.cpu_values.append(cpu_percent)
                self.peak_metrics['cpu'] = max(self.peak_metrics['cpu'], cpu_percent)
                self.peak_metrics['cpu_low'] = min(self.peak_metrics['cpu_low'], cpu_percent)
            
            # Get memory and I/O stats
            memory = psutil.Process().memory_info()
            current_io = psutil.disk_io_counters()
            
            # Calculate disk I/O deltas
            read_delta = current_io.read_bytes - self.last_io.read_bytes
            write_delta = current_io.write_bytes - self.last_io.write_bytes
            
            # Handle counter resets
            if read_delta < 0:
                read_delta = 0
            if write_delta < 0:
                write_delta = 0
                
            # Update totals
            self.total_read_bytes += read_delta
            self.total_write_bytes += write_delta
            
            # Convert to MB
            total_read_mb = self.total_read_bytes / (1024 * 1024)
            total_write_mb = self.total_write_bytes / (1024 * 1024)
            
            # Store current I/O values for next iteration
            self.last_io = current_io
            
            metrics = {
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'cpu_percent': round(cpu_percent, 2),
                'cpu_peak': round(self.peak_metrics['cpu'], 2),
                'cpu_low': round(self.peak_metrics['cpu_low'], 2),
                'cpu_avg': round(statistics.mean(self.cpu_values), 2) if self.cpu_values else 0,
                'memory_rss': round(memory.rss / (1024 * 1024), 2),
                'disk_read_mb': round(total_read_mb, 2),
                'disk_write_mb': round(total_write_mb, 2),
                'disk_read_count': current_io.read_count,
                'disk_write_count': current_io.write_count
            }
            
            return metrics
            
        except Exception as e:
            print(f"Error collecting metrics: {e}")
            return None
    
    def monitor(self):
        """Monitor system continuously"""
        while self.monitoring:
            metrics = self.collect_metrics()
            self.metrics.append(metrics)
            time.sleep(self.interval)
    
    def start(self):
        """Start monitoring"""
        self.monitoring = True
        self.start_time = datetime.now()
        self.monitor_thread = threading.Thread(target=self.monitor)
        self.monitor_thread.start()
        print("Monitoring started...")
    
    def stop(self):
        """Stop monitoring and save results"""
        self.monitoring = False
        if hasattr(self, 'monitor_thread'):
            self.monitor_thread.join()
        
        # Create results directory if it doesn't exist
        results_dir = Path('monitoring_results')
        results_dir.mkdir(exist_ok=True)
        
        # Save metrics to CSV
        filename = results_dir / f'metrics_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        with open(filename, 'w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=self.metrics[0].keys())
            writer.writeheader()
            writer.writerows(self.metrics)
        
        print(f"Monitoring stopped. Results saved to {filename}")
        self.analyze_results()
    
    def analyze_results(self):
        """Analyze collected metrics"""
        if not self.metrics:
            print("No metrics collected")
            return
        
        # Calculate statistics
        cpu_values = [m['cpu_percent'] for m in self.metrics]
        memory_values = [m['memory_rss'] for m in self.metrics]
        disk_read = [m['disk_read_mb'] for m in self.metrics]
        disk_write = [m['disk_write_mb'] for m in self.metrics]
        
        analysis = {
            'duration': f"{(datetime.now() - self.start_time).total_seconds():.2f} seconds",
            'cpu_avg': f"{sum(cpu_values)/len(cpu_values):.2f}%",
            'cpu_max': f"{max(cpu_values):.2f}%",
            'memory_avg': f"{sum(memory_values)/len(memory_values):.2f} MB",
            'memory_max': f"{max(memory_values)::.2f} MB",
            'disk_read_total': f"{max(disk_read):.2f} MB",
            'disk_write_total': f"{max(disk_write):.2f} MB"
        }
        
        print("\nPerformance Analysis:")
        print("-" * 50)
        for key, value in analysis.items():
            print(f"{key.replace('_', ' ').title()}: {value}")

    def analyze_performance(self):
        """Analyze collected metrics and generate insights"""
        if not self.metrics:
            return "No metrics collected"

        cpu_values = [m['cpu_percent'] for m in self.metrics]
        memory_values = [m['memory_rss'] for m in self.metrics]
        disk_read = [m['disk_read_mb'] for m in self.metrics]
        disk_write = [m['disk_write_mb'] for m in self.metrics]

        analysis = {
            'General': {
                'Monitoring Duration': f"{(datetime.now() - self.start_time).total_seconds():.2f} seconds",
                'Total Samples': len(self.metrics)
            },
            'CPU Analysis': {
                'Average CPU Usage': f"{statistics.mean(cpu_values):.2f}%",
                'Peak CPU Usage': f"{max(cpu_values):.2f}%",
                'CPU Usage Variance': f"{statistics.variance(cpu_values):.2f}"
            },
            'Memory Analysis': {
                'Average Memory Usage': f"{statistics.mean(memory_values):.2f} MB",
                'Peak Memory Usage': f"{max(memory_values):.2f} MB",
                'Memory Growth Rate': f"{(memory_values[-1] - memory_values[0]) / len(memory_values):.2f} MB/sample"
            },
            'Disk I/O Analysis': {
                'Total Read': f"{max(disk_read):.2f} MB",
                'Total Write': f"{max(disk_write):.2f} MB",
                'I/O Operations': f"{sum([m['disk_read_count'] + m['disk_write_count'] for m in self.metrics])}"
            }
        }
        
        return analysis

    def analyze_metrics(self):
        """Analyze collected metrics and generate insights"""
        try:
            if not self.metrics:
                return {}

            # Calculate statistics using statistics module
            cpu_values = [m['cpu_percent'] for m in self.metrics]
            memory_values = [m['memory_rss'] for m in self.metrics]
            disk_read = [m['disk_read_mb'] for m in self.metrics]
            disk_write = [m['disk_write_mb'] for m in self.metrics]

            monitoring_duration = (datetime.now() - self.start_time).total_seconds()
            memory_growth = (memory_values[-1] - memory_values[0]) / len(memory_values) if memory_values else 0

            analysis = {
                'CPU Analysis': {
                    'Average CPU Usage': f"{statistics.mean(cpu_values):.2f}%",
                    'CPU Usage Variance': f"{statistics.variance(cpu_values) if len(cpu_values) > 1 else 0:.2f}",
                    'Peak CPU Usage': f"{max(cpu_values):.2f}%"
                },
                'Memory Analysis': {
                    'Average Memory Usage': f"{statistics.mean(memory_values):.2f} MB",
                    'Memory Growth Rate': f"{memory_growth:.2f} MB/sample",
                    'Total Read': f"{self.io_counters.read_bytes / (1024 * 1024):.2f} MB",
                    'Total Write': f"{self.io_counters.write_bytes / (1024 * 1024)::.2f} MB"
                },
                'General': {
                    'Monitoring Duration': f"{monitoring_duration:.2f} seconds",
                    'Total Samples': f"{len(self.metrics)}"
                }
            }
            
            return analysis
            
        except Exception as e:
            print(f"Error in analyze_metrics: {e}")
            return {}

    def export_to_excel(self, wb):
        """Export monitoring data to Excel"""
        try:
            # Create monitoring sheet
            monitor_ws = wb.create_sheet(title="System Monitoring")
            current_row = 1

            # Get analysis data
            analysis = self.analyze_performance()
            
            # Add analysis sections
            for section, metrics in analysis.items():
                # Add section header
                cell = monitor_ws.cell(row=current_row, column=1, value=section)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="86C232", end_color="86C232", fill_type="solid")
                current_row += 1
                
                # Add metrics
                for metric, value in metrics.items():
                    monitor_ws.cell(row=current_row, column=1, value=metric)
                    monitor_ws.cell(row=current_row, column=2, value=str(value))
                    current_row += 1
                current_row += 1

            # Auto-adjust columns
            for column_cells in monitor_ws.columns:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                monitor_ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)
            
            return wb
            
        except Exception as e:
            print(f"Error in export_to_excel: {e}")
            return wb