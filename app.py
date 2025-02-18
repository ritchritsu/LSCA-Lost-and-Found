from cs50 import SQL
from flask import Flask, flash, redirect, render_template, request, session, jsonify, url_for, send_file
from flask_session import Session
import hashlib
import os
import base64
from hmac import compare_digest
from helpers import login_required
from itsdangerous import URLSafeTimedSerializer
from datetime import datetime
from flask_mail import Mail, Message
import re
from dotenv import load_dotenv
from sentence_transformers import SentenceTransformer
import torch
from sklearn.metrics.pairwise import cosine_similarity
import numpy as np
import time
from functools import lru_cache, wraps
import requests.exceptions
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter
from monitor import SystemMonitor
import atexit
import psutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from monitor import SystemMonitor
import atexit
import psutil

def generate_password_hash(password, method='sha256', salt_length=16):
    """Generate a secure password hash"""
    try:
        # Generate a random salt
        salt = os.urandom(salt_length)
        
        # Convert the password to bytes
        password_bytes = password.encode('utf-8')
        
        # Combine salt and password
        salted_password = salt + password_bytes
        
        # Create hash
        hash_obj = hashlib.sha256(salted_password)
        hash_value = hash_obj.digest()
        
        # Encode salt and hash for storage
        salt_b64 = base64.b64encode(salt).decode('utf-8')
        hash_b64 = base64.b64encode(hash_value).decode('utf-8')
        
        # Return format: method$salt$hash
        return f"sha256${salt_b64}${hash_b64}"
    except Exception as e:
        print(f"Error generating password hash: {e}")
        return None

def check_password_hash(pwhash, password):
    """Verify a password against a hash"""
    try:
        # Split the stored hash into its components
        method, salt_b64, hash_b64 = pwhash.split('$')
        
        # Decode the stored salt and hash
        salt = base64.b64decode(salt_b64)
        stored_hash = base64.b64decode(hash_b64)
        
        # Hash the provided password with the same salt
        password_bytes = password.encode('utf-8')
        salted_password = salt + password_bytes
        hash_obj = hashlib.sha256(salted_password)
        calculated_hash = hash_obj.digest()
        
        # Compare the hashes using a secure comparison
        return compare_digest(stored_hash, calculated_hash)
    except Exception as e:
        print(f"Error checking password hash: {e}")
        return False

load_dotenv()

device = "cpu"  # Force CPU usage
model = SentenceTransformer("all-MiniLM-L6-v2", device=device)
print(f"Using device: {device}")

# Configure application
app = Flask(__name__)

# Add security configurations
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'your-secret-key-here')
app.config['SECURITY_PASSWORD_SALT'] = os.getenv('SECURITY_PASSWORD_SALT', 'your-security-salt-here')

# Ensure templates auto-reload
app.config['TEMPLATES_AUTO_RELOAD'] = True

# Mail configuration
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 465
app.config['MAIL_USE_TLS'] = False
app.config['MAIL_USE_SSL'] = True
app.config['MAIL_USERNAME'] = os.getenv('EMAIL_USER')
app.config['MAIL_PASSWORD'] = os.getenv('EMAIL_PASSWORD')
app.config['MAIL_DEFAULT_SENDER'] = os.getenv('EMAIL_USER')

mail = Mail(app)

# Configure session to use filesystem
app.config["SESSION_PERMANENT"] = False
app.config["SESSION_TYPE"] = "filesystem"
Session(app)

# Configure CS50 Library to use SQLite database
db = SQL("sqlite:///app.db")

def init_db():
    """Initialize database tables"""
    db.execute("""
        CREATE TABLE IF NOT EXISTS audit_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_email TEXT NOT NULL,
            action_type TEXT NOT NULL,
            item_id INTEGER,
            details TEXT,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """)

@app.after_request
def after_request(response):
    """Ensure responses aren't cached"""
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Expires"] = 0
    response.headers["Pragma"] = "no-cache"
    return response

def is_admin():
    """Check if current user is admin"""
    try:
        user_email = db.execute("SELECT email FROM users WHERE id = ?", session["user_id"])[0]["email"]
        return user_email == "ritchangelo.dacanay@lsca.edu.ph"
    except:
        return False

@app.context_processor
def utility_processor():
    """Make functions available to templates"""
    return {
        "is_admin": is_admin
    }

@app.route("/")
@login_required
def index():
    """Show submission form or admin dashboard based on user"""
    if session.pop("reset_email_sent", False):
        flash("Password reset instructions have been sent to your email. Please check your email.")
    try:
        user_email = db.execute("SELECT email FROM users WHERE id = ?", session["user_id"])[0]["email"]
        
        if user_email == "ritchangelo.dacanay@lsca.edu.ph":
            items = db.execute("SELECT * FROM items")
            return render_template("admin-dashboard.html", items=items)
        else:
            return render_template("submission.html")
    
    except Exception as e:
        print(f"Error fetching items: {e}")
        return render_template("admin-dashboard.html", items=[])
    



@app.route("/login", methods=["GET", "POST"])
def login():
    """Log user in"""
    session.clear()

    if request.method == "POST":
        email = request.form.get("email")
        password = request.form.get("password")

        if not email or not email.endswith("@lsca.edu.ph"):
            flash("Invalid email. Please use your LSCA email.")
            return render_template("login.html")

        try:
            rows = db.execute("SELECT * FROM users WHERE email = ?", email)
        except Exception as e:
            flash("An error occurred while checking your credentials.")
            print(f"Error fetching user: {e}")
            return render_template("login.html")

        if not rows:
            flash("No account found with that email address. Please register first.")
            return redirect(url_for('register'))

        if not password:
            flash("Please provide a password")
            return render_template("login.html")

        if len(rows) != 1 or not check_password_hash(rows[0]["hash"], password):
            flash("Invalid email or password")
            return render_template("login.html")

        session["user_id"] = rows[0]["id"]
        flash("Logged in successfully!")
        return redirect("/")

    return render_template("login.html")

@app.route("/logout")
def logout():
    """Log user out"""
    session.clear()
    flash("You have been logged out.")
    return redirect("/")

@app.route("/register", methods=["GET", "POST"])
def register():
    """Register user"""
    if request.method == "POST":
        email = request.form.get("email")
        password = request.form.get("password")
        confirmation = request.form.get("confirmation")

        if not email or not re.match(r"[^@]+@lsca\.edu\.ph$", email):
            flash("Invalid email format. Please use your LSCA email.")
            return render_template("register.html")

        if not password or len(password) < 8:
            flash("Password must be at least 8 characters long")
            return render_template("register.html")

        if password != confirmation:
            flash("Password and confirmation do not match")
            return render_template("register.html")

        try:
            # Check if email already exists
            rows = db.execute("SELECT * FROM users WHERE email = ?", email)
            
            if rows and len(rows) > 0:
                flash("This email is already registered")
                return render_template("register.html")

            # Generate password hash
            hashed_password = generate_password_hash(password)
            if not hashed_password:
                flash("Error during registration. Please try again.")
                return render_template("register.html")

            # Insert new user
            user_id = db.execute(
                "INSERT INTO users (email, hash, is_confirmed) VALUES (?, ?, FALSE)", 
                email, hashed_password
            )
            
            session["user_id"] = user_id
            send_confirmation_email(email)
            
            flash("Registration successful! Please check your email to confirm your account.")
            return redirect(url_for('unconfirmed'))
            
        except Exception as e:
            print(f"Error during registration: {e}")
            flash("An error occurred during registration")
            return render_template("register.html")

    return render_template("register.html")

@app.route("/submit", methods=["GET", "POST"])
@login_required
def submit_item():
    if request.method == "POST":
        try:
            # Get form data
            item_status = request.form.get("item_status")
            date = request.form.get("date")
            item_description = request.form.get("item_description")
            lost_location = request.form.get("lost_location")
            found_location = request.form.get("found_location")
            email = request.form.get("email")
            grade_and_section = request.form.get("grade_and_section")
            phone_number = request.form.get("phone_number")
            image_data = request.form.get("image_data")

            # Validate required fields
            if not all([item_status, date, item_description, email, grade_and_section, phone_number]):
                flash("Please fill in all required fields")
                return redirect("/submit")

            # Set correct location based on status
            if item_status.lower() == "lost":
                if not lost_location:
                    flash("Please provide the lost location")
                    return redirect("/submit")
                location = lost_location
                found_location = None
            else:  # Found
                if not found_location:
                    flash("Please provide the found location")
                    return redirect("/submit")
                location = found_location
                lost_location = None

            # Insert item
            result = db.execute("""
                INSERT INTO items 
                (item_status, lost_date, found_date, item_description, 
                 location, found_location, email, grade_and_section, 
                 phone_number, image_url) 
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
                item_status,
                date if item_status.lower() == "lost" else None,
                date if item_status.lower() == "found" else None,
                item_description,
                location,
                found_location,
                email,
                grade_and_section,
                phone_number,
                image_data
            )

            # Get the last inserted ID
            item_id = db.execute("SELECT last_insert_rowid()")[0]['last_insert_rowid()']

            # Log the action
            user_email = db.execute("SELECT email FROM users WHERE id = ?", 
                                  session["user_id"])[0]["email"]
            
            log_action(
                user_email=user_email,
                action_type="item_submission",
                item_id=item_id,
                details=f"Submitted {item_status} item: {item_description}"
            )

            flash("Your item has been submitted successfully!")
            return redirect("/")

        except Exception as e:
            print(f"Error inserting item: {e}")
            flash("An error occurred while submitting your item.")
            return redirect("/submit")

    return render_template("submission.html")

@app.route("/found")
@login_required
def found_items():
    """Show all found items"""
    items = db.execute("SELECT * FROM items WHERE item_status = 'found'")
    return render_template("found-items.html", items=items)

@app.route("/lost")
@login_required
def lost_items():
    """Show all lost items"""
    items = db.execute("SELECT * FROM items WHERE item_status = 'lost'")
    return render_template("lost-items.html", items=items)

@app.route("/update-table-data", methods=["POST"])
@login_required
def update_table_data():
    """Update the table data in the database"""
    try:
        data = request.json
        valid_fields = ["item_status", "lost_date", "found_date", "item_description", 
                       "location", "found_location", "email", "grade_and_section", 
                       "phone_number", "image_url"]

        for entry in data['data']:
            id = entry['id']
            field = entry['field']
            value = entry['value']

            if field not in valid_fields:
                return jsonify({'success': False, 'error': f"Invalid field: {field}"})

            query = f"UPDATE items SET {field} = ? WHERE id = ?"
            db.execute(query, value, id)

        return jsonify({'success': True})
    
    except Exception as e:
        print(f"Error updating data: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route("/update-status", methods=["POST"])
@login_required
def update_status():
    """Update the item status in the database"""
    try:
        data = request.json
        item_id = data.get("id")
        item_status = data.get("item_status")

        if not item_id or not item_status:
            return jsonify({'success': False, 'error': 'Invalid request'})

        db.execute("UPDATE items SET item_status = ? WHERE id = ?", item_status, item_id)
        return jsonify({'success': True})
    
    except Exception as e:
        print(f"Error updating status: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        email = request.form.get("email")
        
        if not email or not email.endswith("@lsca.edu.ph"):
            flash("Invalid email. Please use your LSCA email.")
            return render_template("forgot-password.html")
            
        user = db.execute("SELECT * FROM users WHERE email = ?", email)
        if not user:
            flash("No account found with that email address.")
            return render_template("forgot-password.html")
            
        try:
            token = generate_token(email)
            reset_url = url_for('reset_password', token=token, _external=True)
            
            html = render_template('reset_password_email.html', reset_url=reset_url)
            subject = "Password Reset Request"
            msg = Message(
                subject,
                recipients=[email],
                html=html,
                sender=app.config['MAIL_DEFAULT_SENDER']
            )
            
            mail.send(msg)
            flash("Password reset instructions have been sent to your email. Please check your email.")
            return render_template("forgot-password.html")
            
        except Exception as e:
            print(f"Error sending email: {e}")
            flash("Error sending reset email. Please try again later.")
            return render_template("forgot-password.html")
            
    return render_template("forgot-password.html")

@app.route("/reset-password/<token>", methods=["GET", "POST"])
def reset_password(token):
    try:
        email = confirm_token(token)
        if not email:
            flash('The password reset link is invalid or has expired.', 'danger')
            return redirect(url_for('login'))
    except:
        flash('The password reset link is invalid or has expired.', 'danger')
        return redirect(url_for('login'))
        
    if request.method == "POST":
        password = request.form.get("password")
        confirmation = request.form.get("confirmation")
        
        if not password or len(password) < 8:
            flash("Password must be at least 8 characters long")
            return render_template("reset-password.html")
            
        if password != confirmation:
            flash("Password and confirmation do not match")
            return render_template("reset-password.html")
            
        try:
            hashed_password = generate_password_hash(password)
            db.execute("UPDATE users SET hash = ? WHERE email = ?", hashed_password, email)
            flash("Your password has been updated. Please login with your new password.")
            return redirect(url_for('login'))
        except Exception as e:
            print(f"Error resetting password: {e}")
            flash("An error occurred. Please try again later.")
            return render_template("reset-password.html")
            
    return render_template("reset-password.html")
def generate_token(email):
    serializer = URLSafeTimedSerializer(app.config['SECRET_KEY'])
    return serializer.dumps(email, salt=app.config['SECURITY_PASSWORD_SALT'])

def confirm_token(token, expiration=3600):
    serializer = URLSafeTimedSerializer(app.config['SECRET_KEY'])
    try:
        email = serializer.loads(
            token,
            salt=app.config['SECURITY_PASSWORD_SALT'],
            max_age=expiration
        )
        return email
    except:
        return False

def send_confirmation_email(user_email):
    token = generate_token(user_email)
    confirm_url = url_for('confirm_email', token=token, _external=True)
    html = render_template('confirm_email.html', confirm_url=confirm_url)
    subject = "Please confirm your email"
    
    msg = Message(
        subject,
        recipients=[user_email],
        html=html,
        sender=app.config['MAIL_DEFAULT_SENDER']
    )
    mail.send(msg)

@app.route('/confirm/<token>')
@login_required
def confirm_email(token):
    try:
        email = confirm_token(token)
    except:
        flash('The confirmation link is invalid or has expired.', 'danger')
        return redirect(url_for('index'))
    
    user_id = session["user_id"]
    user = db.execute("SELECT * FROM users WHERE id = ?", user_id)[0]
    
    if user["is_confirmed"]:
        flash('Account already confirmed.', 'success')
    else:
        db.execute(
            "UPDATE users SET is_confirmed = TRUE, confirmed_on = ? WHERE id = ?",
            datetime.now(), user_id
        )
        flash('You have confirmed your account. Thanks!', 'success')
    
    return redirect(url_for('index'))

@app.route('/resend')
@login_required
def resend_confirmation():
    user_id = session["user_id"]
    user = db.execute("SELECT * FROM users WHERE id = ?", user_id)[0]
    
    if user["is_confirmed"]:
        flash('Your account is already confirmed.', 'success')
        return redirect(url_for('index'))
    
    send_confirmation_email(user["email"])
    flash('A new confirmation email has been sent.', 'success')
    return redirect(url_for('index'))

def check_confirmed(func):
    @wraps(func)
    def decorated_function(*args, **kwargs):
        user_id = session.get("user_id")
        if user_id:
            user = db.execute("SELECT is_confirmed FROM users WHERE id = ?", user_id)[0]
            if not user["is_confirmed"]:
                flash('Please confirm your account!', 'warning')
                return redirect(url_for('unconfirmed'))
        return func(*args, **kwargs)
    return decorated_function

@app.route('/unconfirmed')
@login_required
def unconfirmed():
    user_id = session.get("user_id")
    if user_id:
        user = db.execute("SELECT is_confirmed FROM users WHERE id = ?", user_id)[0]
        if user["is_confirmed"]:
            return redirect(url_for('index'))
    return render_template('unconfirmed.html')

# Initialize model with longer timeout and retries
@lru_cache(maxsize=1)
def get_model():
    cache_dir = "./model_cache"
    os.makedirs(cache_dir, exist_ok=True)
    return SentenceTransformer(
        'all-MiniLM-L6-v2',
        device='cpu',
        cache_folder=cache_dir
    )

# Initialize model lazily
model = None

@app.route("/find-similar-items", methods=["POST"])
@login_required
def find_similar_items():
    global model
    
    try:
        # Initialize model on first request
        if model is None:
            try:
                model = get_model()
            except Exception as e:
                print(f"Error loading model: {e}")
                return jsonify({
                    'success': False, 
                    'error': 'Model initialization failed. Please try again later.'
                })

        data = request.json
        description = data.get("description")
        
        if not description:
            return jsonify({
                'success': False,
                'error': 'No description provided'
            })

        # Get all found items
        found_items = db.execute("SELECT * FROM items WHERE item_status = 'Found'")
        
        if not found_items:
            return jsonify({'success': True, 'items': []})
        
        # Create embeddings
        query_embedding = model.encode([description])[0]
        found_descriptions = [item['item_description'] for item in found_items]
        found_embeddings = model.encode(found_descriptions)
        
        # Calculate similarities
        similarities = cosine_similarity([query_embedding], found_embeddings)[0]
        
        # Sort items by similarity
        similar_items = []
        for idx, similarity in enumerate(similarities):
            if (similarity > 0.3):  # Similarity threshold
                item = found_items[idx]
                item['similarity'] = float(similarity)
                similar_items.append(item)
        
        # Sort by similarity score
        similar_items.sort(key=lambda x: x['similarity'], reverse=True)
        
        return jsonify({'success': True, 'items': similar_items[:5]})  # Return top 5 matches
        
    except Exception as e:
        print(f"Error finding similar items: {e}")
        return jsonify({'success': False, 'error': str(e)})

@app.route("/delete-item", methods=["POST"])
@login_required
def delete_item():
    try:
        data = request.json
        item_id = data.get("id")
        if not item_id:
            return jsonify({"success": False, "error": "No item ID provided."}), 400

        # Get item details before deletion
        item = db.execute("SELECT item_description, item_status FROM items WHERE id = ?", item_id)[0]
        
        db.execute("DELETE FROM items WHERE id = ?", item_id)
        
        # Log the deletion
        log_action(
            "item_deletion",
            f"Deleted {item['item_status']} item: {item['item_description']}",
            item_id
        )
        
        return jsonify({"success": True}), 200
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/download-excel", methods=["GET"])
@login_required
def download_excel():
    """Generate and download Excel backup with all data"""
    if not is_admin():
        return jsonify({"success": False, "error": "Unauthorized"}), 403
    
    try:
        # Create workbook
        wb = openpyxl.Workbook()
        
        # 1. Items Sheet (Admin Dashboard)
        items_ws = wb.active
        items_ws.title = "Items"
        
        # Query items
        items = db.execute("""
            SELECT 
                id, item_status, lost_date, found_date, 
                item_description, location, found_location,
                email, grade_and_section, phone_number
            FROM items ORDER BY id DESC
        """)
        
        if items:
            # Add headers
            headers = list(items[0].keys())
            for col_num, header in enumerate(headers, 1):
                cell = items_ws.cell(row=1, column=col_num, value=header.upper())
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="86C232", end_color="86C232", fill_type="solid")
            
            # Add data
            for row_num, item in enumerate(items, 2):
                for col_num, header in enumerate(headers, 1):
                    items_ws.cell(row=row_num, column=col_num, value=item[header])

        # 2. Audit Logs Sheet
        audit_ws = wb.create_sheet(title="Audit Logs")
        
        # Query audit logs
        audit_logs = db.execute("""
            SELECT 
                id, user_email, action_type, item_id, details,
                datetime(timestamp, 'localtime') as timestamp
            FROM audit_logs ORDER BY timestamp DESC
        """)
        
        if audit_logs:
            # Add headers
            audit_headers = ["ID", "User", "Action", "Item ID", "Details", "Timestamp"]
            for col_num, header in enumerate(audit_headers, 1):
                cell = audit_ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="86C232", end_color="86C232", fill_type="solid")
            
            # Add data
            for row_num, log in enumerate(audit_logs, 2):
                audit_ws.cell(row=row_num, column=1, value=log['id'])
                audit_ws.cell(row=row_num, column=2, value=log['user_email'])
                audit_ws.cell(row=row_num, column=3, value=log['action_type'])
                audit_ws.cell(row=row_num, column=4, value=log['item_id'])
                audit_ws.cell(row=row_num, column=5, value=log['details'])
                audit_ws.cell(row=row_num, column=6, value=log['timestamp'])

        # 3. Add monitoring data and analysis
        if hasattr(app, 'monitor'):
            wb = app.monitor.export_to_excel(wb)

        # Auto-adjust column widths for all sheets
        for ws in wb.worksheets:
            for column_cells in ws.columns:
                length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
                adjusted_width = min(length + 2, 50)  # Cap width at 50
                ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

        # Save to memory buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        filename = f"lost_and_found_backup_{timestamp}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print(f"Error generating Excel: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/audit-log")
@login_required
def audit_log():
    """Show audit log"""
    if not is_admin():
        return redirect("/")
        
    logs = db.execute("""
        SELECT 
            audit_logs.*,
            datetime(timestamp, 'localtime') as local_time
        FROM audit_logs 
        ORDER BY timestamp DESC
    """)
    return render_template("audit-log.html", logs=logs)

@app.route("/audit-logs")
@login_required
def audit_logs():
    """Display audit logs"""
    # Ensure admin
    if not is_admin():
        flash("Access denied. Admin only.")
        return redirect("/")

    try:
        # Create table if it doesn't exist
        db.execute("""
            CREATE TABLE IF NOT EXISTS audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_email TEXT NOT NULL,
                action_type TEXT NOT NULL,
                item_id INTEGER,
                details TEXT,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # Fetch audit logs
        logs = db.execute("""
            SELECT
                id,
                user_email,
                action_type,
                item_id,
                details,
                datetime(timestamp, 'localtime') as local_time
            FROM audit_logs
            ORDER BY timestamp DESC
        """)

        # Pass logs to audit-log.html
        return render_template("audit-log.html", logs=logs)
    except Exception as e:
        print(f"Error fetching audit logs: {e}")
        flash("Error loading audit logs.")
        return redirect("/")

def log_action(user_email, action_type, item_id, details):
    """Log an action to audit_logs"""
    try:
        db.execute("""
            INSERT INTO audit_logs 
            (user_email, action_type, item_id, details)
            VALUES (?, ?, ?, ?)
        """, user_email, action_type, item_id, details)
        return True
    except Exception as e:
        print(f"Error logging action: {e}")
        return False

# Add this with your other routes

@app.route("/system-monitor")
@login_required
def system_monitor():
    """Show system monitoring dashboard"""
    if not is_admin():
        return redirect("/")
        
    try:
        # First verify table exists
        db.execute("""
            CREATE TABLE IF NOT EXISTS audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_email TEXT NOT NULL,
                action_type TEXT NOT NULL,
                item_id INTEGER,
                details TEXT,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # Get audit logs with formatted timestamp
        logs = db.execute("""
            SELECT 
                audit_logs.id,
                audit_logs.user_email,
                audit_logs.action_type,
                audit_logs.item_id,
                audit_logs.details,
                datetime(audit_logs.timestamp, 'localtime') as local_time
            FROM audit_logs 
            ORDER BY timestamp DESC
            LIMIT 100
        """)
        
        # Debug print
        print(f"Found {len(logs)} audit log entries")
        
        # Verify we have all required fields
        for log in logs:
            log['local_time'] = log['local_time'] or 'N/A'
            log['user_email'] = log['user_email'] or 'N/A'
            log['action_type'] = log['action_type'] or 'N/A'
            log['item_id'] = log['item_id'] or 'N/A'
            log['details'] = log['details'] or 'N/A'

        # Pass logs and additional context
        return render_template(
            "system-monitor.html",
            logs=logs,
            log_count=len(logs),
            active_page='system_monitor'
        )
        
    except Exception as e:
        print(f"Error fetching audit logs: {e}")
        # Return empty list rather than redirecting
        return render_template("system-monitor.html", logs=[], log_count=0)

@app.route("/system-metrics")
@login_required
def system_metrics():
    """Get current system metrics"""
    if not is_admin():
        return jsonify({"error": "Unauthorized"}), 403
        
    try:
        metrics = app.monitor.collect_metrics()
        metrics['analysis'] = app.monitor.analyze_performance()
        metrics['peak_metrics'] = app.monitor.peak_metrics
        return jsonify(metrics)
    except Exception as e:
        print(f"Error getting metrics: {e}")
        return jsonify({"error": str(e)}), 500

# Initialize monitor
monitor = SystemMonitor(interval=1)

# Start monitoring when app starts
@app.before_first_request
def start_monitoring():
    monitor.start()

# Stop monitoring when app stops
atexit.register(lambda: monitor.stop())

# After creating Flask app
app.monitor = SystemMonitor()
app.monitor.start()

if __name__ == "__main__":
    # Pre-load model before running server
    try:
        print("Loading model...")
        model = get_model()
        print("Model loaded successfully")
    except Exception as e:
        print(f"Warning: Model failed to load: {e}")
        print("Model will be loaded on first request")
    
    # Initialize database tables
    init_db()
    
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)